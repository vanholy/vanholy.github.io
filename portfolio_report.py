#!/usr/bin/env python3
"""Portfolio Price Updater - portafoglio_recap_finale.md (Aprile 2026)"""

import yfinance as yf
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import warnings, sys

warnings.filterwarnings("ignore")

TODAY = date.today()
REPORT_DATE = TODAY.strftime("%Y-%m-%d")
REF_DATE = date(2026, 4, 23)

print(f"Report date: {REPORT_DATE}")
print(f"Reference date: {REF_DATE}")

# ─── Portfolio dati da portafoglio_recap_finale.md ─────────────────────────
# "Quote" = numero di quote (unità) detenute (formato italiano: . = migliaia, , = decimale)
# ICHN: 1.954 in italiano = 1954 unità

INSTRUMENTS = {
    "equity": [
        {
            "name": "Vanguard FTSE All-World UCITS ETF (Acc)",
            "label": "VWCE",
            "tickers": ["VWCE.DE", "VWCE.AS", "VWCE.L"],
            "isin": "IE00B3RBWM25",
            "units": 522.77,
            "ref_value": 79968,
        },
        {
            "name": "Vanguard FTSE Developed Europe UCITS ETF (Acc)",
            "label": "VWCG",
            "tickers": ["VWCG.L", "VWCG.DE", "VWCG.MI", "VWCG.AS"],
            "isin": "IE00B945VV12",
            "units": 350.8,
            "ref_value": 19666,
        },
        {
            "name": "iShares Core MSCI EM IMI UCITS ETF (Acc)",
            "label": "EIMI",
            "tickers": ["EIMI.L", "EMIM.L", "EIMI.DE", "EMIM.DE"],
            "isin": "IE00BKM4GZ66",
            "units": 229,
            "ref_value": 10024,
        },
        {
            "name": "iShares MSCI China UCITS ETF USD (Acc)",
            "label": "ICHN",
            "tickers": ["ICHN.L", "ICHN.DE", "ICHN.MI"],
            "isin": "IE00B3JYPS56",
            "units": 1954,
            "ref_value": 9832,
        },
    ],
    "bond": [
        {
            "name": "Liquidità Fineco",
            "label": "CASH",
            "tickers": [],
            "isin": None,
            "units": None,
            "ref_value": 20000,
            "fixed_value": 20000,
            "note": "Cash parcheggiato - valore fisso",
        },
        {
            "name": "Amundi Euro Govt Bond 3-5Y UCITS ETF (Acc)",
            "label": "MTB",
            "tickers": ["CB35.PA", "MTB.MI", "EMTB.L", "C035.PA"],
            "isin": "LU1681045370",
            "units": None,
            "ref_value": 19876,
            "note": "Duration ~3,5 anni - variazione via ratio prezzo",
        },
        {
            "name": "BTP Valore MZ32",
            "label": "BTP-MZ32",
            "tickers": [],
            "isin": "IT0005094088",
            "units": None,
            "ref_value": 10000,
            "fixed_value": 10000,
            "note": "Scad. 01/03/2032 - Prezzo fisso 100 (istruzione utente)",
        },
        {
            "name": "XEON - Xtrackers EUR Overnight Rate Swap",
            "label": "XEON",
            "tickers": ["XEON.DE", "XEON.L", "XEON.MI"],
            "isin": "LU0290358497",
            "units": None,
            "ref_value": 10055,
            "note": "Monetario overnight ~2,4% - variazione via ratio prezzo",
        },
    ],
    "alternative": [
        {
            "name": "iMGP DBi Managed Futures Fund R EUR UCITS ETF",
            "label": "DBMFE",
            "tickers": ["DBMFE.L", "DBMFE.DE", "DBMFE.MI"],
            "isin": "LU2951555585",
            "units": 170,
            "ref_value": 20096,
        },
    ],
}

# ─── Price fetching ─────────────────────────────────────────────────────────

def get_current_price(tickers):
    for t in tickers:
        try:
            tk = yf.Ticker(t)
            hist = tk.history(period="5d", interval="1d")
            if not hist.empty:
                price = float(hist["Close"].iloc[-1])
                if price > 0:
                    try:
                        currency = tk.fast_info.currency
                    except:
                        currency = "EUR"
                    print(f"  [{t}] prezzo: {price:.4f} {currency}")
                    return t, price, currency
        except Exception as e:
            pass
    return None, None, None


def get_historical_price(tickers, target_date):
    for t in tickers:
        try:
            tk = yf.Ticker(t)
            start = (target_date - timedelta(days=10)).strftime("%Y-%m-%d")
            end = (target_date + timedelta(days=10)).strftime("%Y-%m-%d")
            hist = tk.history(start=start, end=end)
            if not hist.empty:
                p = float(hist["Close"].iloc[-1])
                print(f"  [{t}] prezzo storico {target_date}: {p:.4f}")
                return p
        except Exception:
            pass
    return None


# ─── Process all instruments ────────────────────────────────────────────────

results = {"equity": [], "bond": [], "alternative": []}
price_log = []

for category, items in INSTRUMENTS.items():
    print(f"\n=== {category.upper()} ===")
    for inst in items:
        r = dict(inst)
        label = inst["label"]

        if "fixed_value" in inst:
            r["current_price"] = None
            r["hist_price"] = None
            r["current_value"] = inst["fixed_value"]
            r["ticker_used"] = "FISSO"
            r["price_change_pct"] = 0.0
            r["price_ok"] = True
            msg = f"{label}: valore fisso €{inst['fixed_value']:,.0f}"

        elif not inst["tickers"]:
            r["current_price"] = None
            r["hist_price"] = None
            r["current_value"] = inst["ref_value"]
            r["ticker_used"] = "N/D"
            r["price_change_pct"] = 0.0
            r["price_ok"] = False
            msg = f"{label}: nessun ticker - valore di riferimento usato"

        else:
            print(f"\n{label} ({inst['isin']}):")
            ticker_used, price, currency = get_current_price(inst["tickers"])
            r["ticker_used"] = ticker_used
            r["current_price"] = price

            if price is None:
                r["hist_price"] = None
                r["current_value"] = inst["ref_value"]
                r["price_change_pct"] = 0.0
                r["price_ok"] = False
                msg = f"{label}: prezzo NON trovato su Yahoo Finance - valore di riferimento (€{inst['ref_value']:,.0f})"
            elif inst.get("units") is not None:
                # ETF con quote note → valore = quote × prezzo live
                r["hist_price"] = None
                r["current_value"] = inst["units"] * price
                r["price_change_pct"] = (r["current_value"] - inst["ref_value"]) / inst["ref_value"] * 100
                r["price_ok"] = True
                msg = (f"{label}: {inst['units']} quote × €{price:.4f} = €{r['current_value']:,.2f} "
                       f"({r['price_change_pct']:+.1f}% vs ref) [{ticker_used}]")
            else:
                # ETF senza quote note → ratio prezzo storico/attuale
                hist_price = get_historical_price(inst["tickers"], REF_DATE)
                r["hist_price"] = hist_price
                if hist_price and hist_price > 0:
                    ratio = price / hist_price
                    r["current_value"] = inst["ref_value"] * ratio
                    r["price_change_pct"] = (ratio - 1) * 100
                    r["price_ok"] = True
                    msg = (f"{label}: rif {REF_DATE}=€{hist_price:.4f}, oggi=€{price:.4f}, "
                           f"ratio={ratio:.4f} → €{r['current_value']:,.2f} "
                           f"({r['price_change_pct']:+.1f}%) [{ticker_used}]")
                else:
                    r["current_value"] = inst["ref_value"]
                    r["price_change_pct"] = 0.0
                    r["price_ok"] = False
                    msg = f"{label}: prezzo storico non trovato - valore di riferimento usato"

        print(f"  → {msg}")
        price_log.append(msg)
        results[category].append(r)

# ─── Calcolo allocazioni ────────────────────────────────────────────────────

total_equity = sum(r["current_value"] for r in results["equity"])
total_bond   = sum(r["current_value"] for r in results["bond"])
total_alt    = sum(r["current_value"] for r in results["alternative"])
total        = total_equity + total_bond + total_alt

pct_eq  = total_equity / total * 100
pct_bo  = total_bond   / total * 100
pct_al  = total_alt    / total * 100

ref_total = sum(r["ref_value"] for cat in results.values() for r in cat)

print(f"\n{'='*60}")
print(f"TOTALE PORTAFOGLIO: €{total:,.2f}  (rif Aprile 2026: €{ref_total:,.2f})")
print(f"Equity:      €{total_equity:,.2f}  = {pct_eq:.2f}%")
print(f"Bond+Mon:    €{total_bond:,.2f}   = {pct_bo:.2f}%")
print(f"Alternative: €{total_alt:,.2f}   = {pct_al:.2f}%")

def check_status(pct, low, high, warn=2.5):
    if pct < low or pct > high:
        return "RED"
    if pct < low + warn or pct > high - warn:
        return "YELLOW"
    return "GREEN"

st_eq = check_status(pct_eq, 45.7, 65.7, warn=3.0)
st_bo = check_status(pct_bo, 25.0, 35.0, warn=2.0)
st_al = check_status(pct_al,  6.3, 12.3, warn=1.5)

STATUS_LABEL = {"GREEN": "✅ OK", "YELLOW": "⚠️ ATTENZIONE", "RED": "🚨 ALERT"}
print(f"Equity alert:      {st_eq}")
print(f"Bond+Mon alert:    {st_bo}")
print(f"Alternative alert: {st_al}")

# ─── Excel ──────────────────────────────────────────────────────────────────

GREEN_F  = PatternFill("solid", fgColor="C6EFCE")
YELLOW_F = PatternFill("solid", fgColor="FFEB9C")
RED_F    = PatternFill("solid", fgColor="FFC7CE")
BLUE_H   = PatternFill("solid", fgColor="1F4E79")
BLUE_M   = PatternFill("solid", fgColor="2E75B6")
BLUE_L   = PatternFill("solid", fgColor="BDD7EE")
TOTAL_F  = PatternFill("solid", fgColor="E2EFDA")

WH_BOLD  = Font(name="Calibri", bold=True, color="FFFFFF")
WH_NORM  = Font(name="Calibri", color="FFFFFF")
BK_BOLD  = Font(name="Calibri", bold=True)
BK_NORM  = Font(name="Calibri")
BK_SM    = Font(name="Calibri", size=9)

STATUS_FILL = {"GREEN": GREEN_F, "YELLOW": YELLOW_F, "RED": RED_F}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = f"Report {REPORT_DATE}"

r = 1

def mrow(ws, r, ncols=10):
    ws.merge_cells(f"A{r}:{get_column_letter(ncols)}{r}")
    return ws[f"A{r}"]

# Title
c = mrow(ws, r)
c.value = f"REPORT PORTAFOGLIO  —  {REPORT_DATE}"
c.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
c.fill = BLUE_H
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r].height = 32
r += 1

c = mrow(ws, r)
c.value = (f"Prezzi live da Yahoo Finance  |  Riferimento documento: Aprile 2026  |"
           f"  Patrimonio aggiornato: €{total:,.2f}")
c.font = Font(name="Calibri", italic=True, size=10, color="595959")
c.alignment = Alignment(horizontal="center")
r += 2

# ── Allocazione Macro ──────────────────────────────────────────
c = mrow(ws, r)
c.value = "ALLOCAZIONE MACRO"
c.font = WH_BOLD
c.fill = BLUE_M
c.alignment = Alignment(horizontal="center")
ws.row_dimensions[r].height = 20
r += 1

HDR = ["Categoria", "Valore Live", "% Attuale", "Target %", "Range Min", "Range Max", "Stato", "Δ vs Target"]
for i, h in enumerate(HDR):
    cell = ws.cell(row=r, column=i+1)
    cell.value = h
    cell.font = WH_BOLD
    cell.fill = BLUE_L
    cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")
r += 1

MACRO = [
    ("Equity  (VWCE · VWCG · EIMI · ICHN)", total_equity, pct_eq, 55.7, 45.7, 65.7, st_eq),
    ("Bond + Monetario + Cash",               total_bond,   pct_bo, 30.0, 25.0, 35.0, st_bo),
    ("Alternative  (DBMFE)",                  total_alt,    pct_al,  9.3,  6.3, 12.3, st_al),
    ("TOTALE PORTAFOGLIO",                    total,       100.0, 100.0, 100.0, 100.0, "GREEN"),
]

for (cat, val, pct, tgt, lo, hi, st) in MACRO:
    fill = TOTAL_F if cat.startswith("TOTALE") else STATUS_FILL[st]
    delta = pct - tgt
    row_data = [cat, val, pct/100, tgt/100, lo/100, hi/100, STATUS_LABEL.get(st, st), delta/100]
    for i, v in enumerate(row_data):
        cell = ws.cell(row=r, column=i+1)
        cell.value = v
        cell.fill = fill
        cell.font = BK_BOLD if i == 0 or cat.startswith("TOTALE") else BK_NORM
        if i == 1: cell.number_format = "€#,##0.00"
        if i in (2, 3, 4, 5, 7): cell.number_format = "0.00%"
        cell.alignment = Alignment(horizontal="left" if i == 0 else "center")
    r += 1

r += 1

# ── Dettaglio per categoria ────────────────────────────────────
CAT_CFG = [
    ("equity",      "EQUITY",                      "VWCE · VWCG · EIMI · ICHN"),
    ("bond",        "BOND + MONETARIO + CASH",      "XEON · MTB · BTP-MZ32 · Cash"),
    ("alternative", "ALTERNATIVE",                  "DBMFE"),
]

DET_HDR = ["Strumento", "Label", "ISIN", "Ticker Yahoo", "Quote", "Prezzo Rif.", "Valore Rif.", "Prezzo Live", "Valore Live", "Var %"]

for cat_key, cat_name, cat_note in CAT_CFG:
    cat_total     = sum(x["current_value"] for x in results[cat_key])
    cat_ref_total = sum(x["ref_value"]     for x in results[cat_key])
    cat_chg       = (cat_total - cat_ref_total) / cat_ref_total * 100

    c = mrow(ws, r)
    c.value = f"{cat_name}  —  {cat_note}  |  Valore: €{cat_total:,.2f}  ({cat_chg:+.1f}% vs rif)"
    c.font = WH_BOLD
    c.fill = BLUE_M
    c.alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[r].height = 20
    r += 1

    for i, h in enumerate(DET_HDR):
        cell = ws.cell(row=r, column=i+1)
        cell.value = h
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill = BLUE_L
        cell.alignment = Alignment(horizontal="center")
    r += 1

    for inst in results[cat_key]:
        chg = inst.get("price_change_pct", 0.0) or 0.0
        ref_price = (inst["ref_value"] / inst["units"]) if inst.get("units") else None

        # row color
        if not inst.get("price_ok", True):
            row_fill = YELLOW_F
        elif chg < -10:
            row_fill = RED_F
        elif abs(chg) > 5:
            row_fill = YELLOW_F
        else:
            row_fill = GREEN_F

        row_vals = [
            inst["name"],
            inst["label"],
            inst.get("isin") or "—",
            inst.get("ticker_used") or "—",
            inst.get("units"),
            ref_price,
            inst["ref_value"],
            inst.get("current_price"),
            inst["current_value"],
            chg / 100 if chg is not None else None,
        ]

        for i, v in enumerate(row_vals):
            cell = ws.cell(row=r, column=i+1)
            cell.value = v if v is not None else "—"
            cell.fill = row_fill
            cell.font = BK_NORM
            cell.alignment = Alignment(horizontal="left" if i == 0 else "center")
            if i == 4 and isinstance(v, float): cell.number_format = "#,##0.00"
            if i in (5, 7) and isinstance(v, float): cell.number_format = "€#,##0.0000"
            if i in (6, 8) and isinstance(v, float): cell.number_format = "€#,##0.00"
            if i == 9 and isinstance(v, float):
                cell.number_format = "+0.00%;-0.00%"
                if v < -0.05: cell.fill = RED_F
                elif v > 0.05: cell.fill = YELLOW_F
        r += 1

    # subtotale
    for i, v in enumerate(["SUBTOTALE", "", "", "", "", "", cat_ref_total, "", cat_total, cat_chg/100]):
        cell = ws.cell(row=r, column=i+1)
        cell.value = v
        cell.font = BK_BOLD
        cell.fill = TOTAL_F
        cell.alignment = Alignment(horizontal="left" if i == 0 else "center")
        if i == 6 or i == 8: cell.number_format = "€#,##0.00"
        if i == 9: cell.number_format = "+0.00%;-0.00%"
    r += 2

# ── Note prezzi ───────────────────────────────────────────────
c = mrow(ws, r)
c.value = "LOG PREZZI"
c.font = WH_BOLD
c.fill = BLUE_M
r += 1
for note in price_log:
    c = mrow(ws, r)
    c.value = f"• {note}"
    c.font = BK_SM
    r += 1

r += 1

# ── Legenda ───────────────────────────────────────────────────
c = mrow(ws, r)
c.value = "LEGENDA COLORI"
c.font = WH_BOLD
c.fill = BLUE_M
r += 1
for fill, text in [
    (GREEN_F,  "VERDE — Allocazione nel range target / variazione prezzo contenuta (<5%)"),
    (YELLOW_F, "GIALLO — Allocazione vicina ai limiti (entro 3 pp) / variazione prezzo significativa (5-10%)"),
    (RED_F,    "ROSSO — Allocazione fuori dai limiti target / variazione prezzo negativa elevata (>10%)"),
]:
    c = mrow(ws, r)
    c.value = text
    c.font = BK_SM
    c.fill = fill
    r += 1

# ── Column widths ─────────────────────────────────────────────
for col, w in zip(range(1, 11), [42, 10, 18, 14, 10, 14, 14, 14, 14, 10]):
    ws.column_dimensions[get_column_letter(col)].width = w

xlsx_path = f"/home/user/vanholy.github.io/report_{REPORT_DATE}.xlsx"
wb.save(xlsx_path)
print(f"\nExcel salvato: {xlsx_path}")

# ─── Markdown ───────────────────────────────────────────────────────────────

SE = {"GREEN": "✅", "YELLOW": "⚠️", "RED": "🚨"}
SL = {"GREEN": "OK", "YELLOW": "ATTENZIONE", "RED": "ALERT"}

md = [
    f"# Report Portafoglio — {REPORT_DATE}",
    "",
    f"*Prezzi live da Yahoo Finance — Riferimento documento: Aprile 2026*",
    "",
    "---",
    "",
    f"## Patrimonio Totale Aggiornato: €{total:,.2f}",
    f"*(Riferimento Aprile 2026: €{ref_total:,.2f} — variazione: {(total-ref_total)/ref_total*100:+.1f}%)*",
    "",
    "---",
    "",
    "## Allocazione Macro",
    "",
    "| Categoria | Valore Live | % Attuale | Target | Range | Stato |",
    "|---|---|---|---|---|---|",
    f"| Equity | €{total_equity:,.2f} | {pct_eq:.1f}% | 55,7% | 45,7%–65,7% | {SE[st_eq]} {SL[st_eq]} |",
    f"| Bond + Monetario + Cash | €{total_bond:,.2f} | {pct_bo:.1f}% | 30,0% | 25%–35% | {SE[st_bo]} {SL[st_bo]} |",
    f"| Alternative (DBMFE) | €{total_alt:,.2f} | {pct_al:.1f}% | 9,3% | 6,3%–12,3% | {SE[st_al]} {SL[st_al]} |",
    f"| **TOTALE** | **€{total:,.2f}** | **100%** | — | — | — |",
    "",
    "---",
    "",
]

alerts = []
if st_eq != "GREEN": alerts.append(f"- {SE[st_eq]} **EQUITY**: {pct_eq:.1f}% — fuori range 45,7%–65,7%" if st_eq=="RED" else f"- {SE[st_eq]} **EQUITY**: {pct_eq:.1f}% — vicino al limite (45,7%–65,7%)")
if st_bo != "GREEN": alerts.append(f"- {SE[st_bo]} **BOND+MONETARIO**: {pct_bo:.1f}% — fuori range 25%–35%" if st_bo=="RED" else f"- {SE[st_bo]} **BOND+MONETARIO**: {pct_bo:.1f}% — vicino al limite (25%–35%)")
if st_al != "GREEN": alerts.append(f"- {SE[st_al]} **ALTERNATIVE**: {pct_al:.1f}% — fuori range 6,3%–12,3%" if st_al=="RED" else f"- {SE[st_al]} **ALTERNATIVE**: {pct_al:.1f}% — vicino al limite (6,3%–12,3%)")

if alerts:
    md += ["## ⚠️ Sbilanciamenti Rilevati", ""] + alerts + ["", "---", ""]
else:
    md += ["## ✅ Nessun Sbilanciamento — Allocazione nella norma", "", "---", ""]

for cat_key, cat_name, cat_note in CAT_CFG:
    cat_total     = sum(x["current_value"] for x in results[cat_key])
    cat_ref_total = sum(x["ref_value"]     for x in results[cat_key])
    cat_chg       = (cat_total - cat_ref_total) / cat_ref_total * 100
    pct_cat       = cat_total / total * 100

    md += [
        f"## {cat_name}",
        f"**Valore: €{cat_total:,.2f}** ({pct_cat:.1f}% portafoglio) | Variazione vs rif: {cat_chg:+.1f}%",
        "",
        "| Stato | Strumento | Label | ISIN | Ticker | Quote | Prezzo Live | Valore Live | Var % | Note |",
        "|---|---|---|---|---|---|---|---|---|---|",
    ]
    for inst in results[cat_key]:
        icon = "✅" if inst.get("price_ok", True) else "⚠️"
        units_s = f"{inst['units']:,.2f}" if inst.get("units") else "—"
        price_s = f"€{inst['current_price']:,.4f}" if inst.get("current_price") else "—"
        chg_s   = f"{inst.get('price_change_pct', 0):+.1f}%" if inst.get("current_price") else "N/D"
        note_s  = inst.get("note", "")
        md.append(
            f"| {icon} | {inst['name']} | {inst['label']} | {inst.get('isin') or '—'} | "
            f"{inst.get('ticker_used') or '—'} | {units_s} | {price_s} | "
            f"€{inst['current_value']:,.2f} | {chg_s} | {note_s} |"
        )
    md += [
        f"| | **SUBTOTALE** | | | | | | **€{cat_total:,.2f}** | **{cat_chg:+.1f}%** | |",
        "", "---", "",
    ]

md += [
    "## Log Prezzi",
    "",
]
for msg in price_log:
    md.append(f"- {msg}")

md += [
    "",
    "---",
    "",
    "## Note Metodologiche",
    "",
    "- **Equity (VWCE, VWCG, EIMI, ICHN)**: `valore = quote × prezzo live Yahoo Finance`",
    "- **DBMFE**: `valore = quote × prezzo live Yahoo Finance`",
    "- **MTB, XEON**: senza quote esplicite → `valore = valore_rif × (prezzo_live / prezzo_rif_apr2026)`",
    "- **BTP IT0005094088**: prezzo fisso 100 per istruzione utente → valore invariato a €10.000",
    "- **Cash Fineco**: valore fisso €20.000",
    "- **Soglie alert**: Equity 45,7%–65,7% | Bond+Mon 25%–35% | Alternative 6,3%–12,3%",
    "- **⚠️ Nota ticker**: il file originale riporta EIMI (non EMIM) e ICHN (non ICGA), e MTB (non EM35/IBGS). Utilizzati i ticker del file.",
    "",
    "---",
    "",
    f"*Report generato automaticamente il {REPORT_DATE}. Non costituisce consulenza finanziaria.*",
]

md_content = "\n".join(md)
md_path = f"/home/user/vanholy.github.io/report_{REPORT_DATE}.md"
with open(md_path, "w", encoding="utf-8") as f:
    f.write(md_content)
print(f"Markdown salvato: {md_path}")

# ─── Output per passi successivi ────────────────────────────────────────────
print(f"\nFILE_XLSX={xlsx_path}")
print(f"FILE_MD={md_path}")
print(f"TOTAL={total:.2f}")
print(f"PCT_EQ={pct_eq:.2f}")
print(f"PCT_BO={pct_bo:.2f}")
print(f"PCT_AL={pct_al:.2f}")
print(f"ST_EQ={st_eq}")
print(f"ST_BO={st_bo}")
print(f"ST_AL={st_al}")
print(f"REF_TOTAL={ref_total:.2f}")
